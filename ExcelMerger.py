import json

from openpyxl import load_workbook
from pyexcel.cookbook import merge_all_to_a_book
import glob

JSON_PATH = "data.json"


class ExcelMerger(object):

    def __init__(self, transaction_spreadsheet, workbook):
        self.path = workbook
        self.workbook_dest = load_workbook(workbook)
        self.ws_dest = self.workbook_dest['Raw Data']

        merge_all_to_a_book(glob.glob("{}/*.csv".format("Transactions")), "{}/output.xlsx".format("Transactions"))
        workbook_source = load_workbook('Transactions/output.xlsx')
        self.ws_source = workbook_source.active

    def process_transactions(self):
        current_row = self._find_max_row()
        for row in self.ws_source.iter_rows():
            if row[0].value == 'Account Type':
                continue
            self.ws_dest['A'+str(current_row)] = self._format_date(row[2].value)
            if row[5].value:
                description = row[4].value + ' ' + row[5].value
            else:
                description = row[4].value
            self.ws_dest['B'+str(current_row)] = description
            self.ws_dest['C'+str(current_row)] = row[6].value
            self.ws_dest['C'+str(current_row)].number_format = u'"$"#,##0.00'
            self.ws_dest['D'+str(current_row)] = row[0].value
            key = self.get_category_key(description)
            sub_category, main_category = self._set_category(key)
            self.ws_dest['J' + str(current_row)] = sub_category
            self.ws_dest['I' + str(current_row)] = main_category
            current_row = current_row + 1
        self.workbook_dest.save(self.path)

    def _find_max_row(self):
        for max_row, row in enumerate(self.ws_dest, 1):
            if all(c.value is None for c in row):
                return max_row
        return max_row

    @staticmethod
    def _format_date(string):
        str_array = string.split("/")
        formatted_date = str_array[1] + "/" + str_array[0] + "/" + str_array[2]
        return formatted_date

    @staticmethod
    def _set_category(key):
        with open(JSON_PATH, "r+") as f:
            category_data = json.load(f)
            if key in category_data["Index"]:
                sub_category = category_data["Index"][key]
                main_category = category_data["SubCategories"][sub_category]
                return sub_category, main_category
            else:
                sub_category = ''
                while sub_category == '' or sub_category == 'list':
                    sub_category = input(f"What is the sub category for {key}? Enter 'list' to view existing categories")
                    if sub_category == 'list':
                        for k, v in category_data["SubCategories"].items():
                            print(k + " : " + v)
                category_data["Index"][key] = sub_category
                if sub_category in category_data["SubCategories"]:
                    main_category = category_data["SubCategories"][sub_category]
                else:
                    is_valid_main_category = False
                    while not is_valid_main_category:
                        main_category = input(f"What is the main category for {sub_category}? "
                                              f"\n Luxury, Income, Necessity or Transfer")
                        if main_category in category_data["MainCategories"]:
                            is_valid_main_category = True
                        else:
                            print(f"{main_category} is an invalid main category")
                    category_data["SubCategories"][sub_category] = main_category
                f.seek(0)
                json.dump(category_data, f, indent=4, separators=(',', ': '), sort_keys=True)
                f.truncate()
                return sub_category, main_category

    @staticmethod
    def has_numbers(input_string):
        return any(char.isdigit() for char in input_string)

    @staticmethod
    def get_category_key(description):
        phrases = description.split(' ')
        key = ""
        for phrase in phrases:
            if not ExcelMerger.has_numbers(phrase):
                key = " ".join([key, phrase])
        return key
